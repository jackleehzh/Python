import openpyxl
import os
import shutil


date = input('请输入排课日期（如，2020-09-17）:')
riqi = date[-5:]
riqi = riqi[:2] + '.' + riqi[3:]
if riqi[0] == '0':
    riqi = riqi[1:]

if riqi[-2] == '0':
    riqi = riqi[:-2] + riqi[-1]
print(riqi)

def newfile(path):
    path=path.strip()
    path=path.rstrip("\\")
    isExists=os.path.exists(path)

    if not isExists:
        os.makedirs(path)
        print(path+' 创建成功')
        return True
    else:
        print(path+' 目录已存在')

    return False


def copyfile(srcdir, dstdir, file_name):
    if os.path.isdir(dstdir):
        srcfile = os.path.join(srcdir, file_name)
        shutil.copy(srcfile, dstdir)


def rename_file(dstdir, file_name):
    new_file_name = date + file_name[10:]
    if new_file_name == date:
        return
    file_path = os.path.join(dstdir, file_name)
    new_file_path = os.path.join(dstdir, new_file_name)

    try:
        os.rename(file_path, new_file_path)
    except Exception as e:
        print(e)
        print('rename file fail\r\n')
    else:
        print('rename file success\r\n')


def read_data():
    wb = openpyxl.load_workbook('./孩子们.xlsx')

    sh = wb['Sheet2']

    all_data = []

    columns = {}
    for cases in list(sh.rows)[:2]:
        c = 0
        for case in cases:
            if case.value == '班型':
                columns['班型'] = c

            if case.value == '带班老师':
                columns['带班老师'] = c
            if case.value is not None and type(case.value) is not int and riqi in case.value:
                columns[riqi] = c

            c += 1

    c = 0
    for cases in list(sh.rows)[2:]:
        line = []
        if cases[columns[riqi]].value is not None:
            line.append(file_pair[c])
            line.append(cases[columns['带班老师']].value)
            line.append(cases[columns[riqi]].value)
            all_data.append(line)

        if cases[columns['班型']].value is not None:
            c += 1

    wb.close()
    return all_data


file_pair = []
file_pair.append('09点40分思维课')
file_pair.append('09点40分启蒙课')
file_pair.append('09点40分Scratch课')
file_pair.append('09点40分Python课')
file_pair.append('14点30分思维课')
file_pair.append('14点30分启蒙课')
file_pair.append('14点30分Scratch课')
file_pair.append('14点30分Python课')
file_pair.append('18点30分思维课')
file_pair.append('19点00分启蒙课')
file_pair.append('18点30分Scratch课')
file_pair.append('18点30分Python课')

datas = read_data()
datas2 = []

for data in datas:
    d = []
    d.append(data[0])
    d.append(data[1])
    d.append(data[2].split())
    datas2.append(d)

for data in datas2:
    print(data)

newpath="./"
newpath += date + '课表'

oldpath = './2020.09.12课表'


newfile(newpath)
for root, dirs, files in os.walk(oldpath):

    for f in files:
        for data in datas2:
            if data[0] in f:
                copyfile(oldpath, newpath, f)
                rename_file(newpath, f)

def writeData(datas2):
    for root, dirs, files in os.walk(newpath):
        for f in files:
            for data in datas2:
                if data[0] in f:

                    keci = 1
                    f2 = f
                    if data[2][-1].find('1') or data[2][-1].find('一'):
                        keci = '1'
                        f2 = f
                    if data[2][-1].find('2') or data[2][-1].find('二'):
                        keci = '2'
                        f2 = f[:-10] + '二' + f[-9:]
                    if data[2][-1].find('3') or data[2][-1].find('三'):
                        keci = '3'
                        f2 = f[:-10] + '三' + f[-9:]
                    if data[2][-1].find('4') or data[2][-1].find('四'):
                        keci = '4'
                        f2 = f[:-10] + '四' + f[-9:]

                    try:
                        os.rename(os.path.join(root, f), os.path.join(root, f2))
                    except Exception as e:
                        print(e)
                        print('rename file fail\r\n')
                    else:
                        print('rename file success\r\n')

                    wb = openpyxl.load_workbook(os.path.join(root, f2))
                    sh = wb['1班']

                    index = sh.cell(row=1, column=1).value.find('1次')
                    sh.cell(row=1, column=1).value = sh.cell(row=1, column=1).value[:index] + keci + sh.cell(row=1, column=1).value[index + 1:]
                    index = sh.cell(row=1, column=1).value.find('-')
                    sh.cell(row=1, column=1).value = sh.cell(row=1, column=1).value[:index - 4] + date + sh.cell(row=1,
                                                                                                             column=1).value[
                                                                                                     index + 6:]
                    print(sh.cell(row=1, column=1).value)
                    c = 3
                    for r in range(len(data[2]) - 1):
                        sh.cell(row=c + r, column=2).value = data[2][r]
                    wb.save(os.path.join(root, f2))
                    wb.close()


writeData(datas2)